#!/usr/bin/env python3
"""
Backtest the shakeout scanner on REAL historical data from Yahoo Finance
(no API token needed - runs locally and inside GitHub Actions).

For every trading day in the window it runs the SAME logic as the live
scanner (BOS -> flush -> SSL hold -> reversal, + 26W-high proximity guard,
+ prefilter conditions) and records each signal with its forward returns:

    entry      = next day's OPEN   (alert at close, enter next open)
    r3/r5/r7/r10/r15 = close[t+k] vs entry
    max15      = best close within 15 sessions vs entry
    min15      = worst close within 15 sessions vs entry
    big_move   = 1 if max15 >= +8% (the Sportking-style pop) else 0

    The summary prints a score-bucket table (50/55/60/65/70/75/80) so you
    can see the win rate and big-move rate climb as the threshold rises -
    the pattern is a SHORT-TERM bounce (3-7 days) and the +8%+ moves only
    fire on the highest-quality setups.

Usage:
    pip install yfinance

    # Time period presets (1 month, 6 months, 1 year, 2 years, 5 years):
    python backtest.py --period 1m --limit 300
    python backtest.py --period 6m --limit 300
    python backtest.py --period 1y --limit 500
    python backtest.py --period 2y --limit 500
    python backtest.py --period 5y --limit 0

    # or an explicit window / score threshold:
    python backtest.py --years 2 --limit 500 --min-score 55

    A period/year window is the ANALYSIS window (signals are only reported
    inside it). The pattern's own lookback (min_bars ~ 26 weeks) is fetched
    automatically on top of it, so short presets like --period 1m still
    have the full BOS/SSL history behind every reportable day.

Output: signals_backtest.csv + signals_backtest.xlsx (Excel with a
        'Signals' sheet containing EVERY signal incl. full score component
        breakdown, and a 'Summary' sheet with win-rate stats + score buckets) and a
        printed win-rate summary, including a score>=70 vs score<70 split.

If most symbols fail to fetch (Yahoo rate-limiting the runner IP), the run
is a DATA OUTAGE: the summary says so loudly and the process exits non-zero
so a CI run turns RED instead of a green, meaningless "0 signals".
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import os
import statistics
import sys
import time

import numpy as np

from config import ScanConfig
from indicators import avg_volume
from pattern import _score
from prefilter import passes_prefilter

# exit code when a run was a data outage (same convention as scanner.py:
# distinct from 1=crash, 2=usage) - a CI run must turn RED, not die green
# with a meaningless "0 signals"
EXIT_DATA_OUTAGE = 3

# ---------------------------------------------------------------------------
# universe (static list of liquid NSE names - the backtest no longer calls
# the Dhan API, so the scrip-master universe is not available here)
# ---------------------------------------------------------------------------
UNIVERSE = [
    "RELIANCE", "TCS", "HDFCBANK", "ICICIBANK", "INFY", "SBIN", "BHARTIARTL",
    "ITC", "LT", "HINDUNILVR", "KOTAKBANK", "AXISBANK", "BAJFINANCE", "MARUTI",
    "SUNPHARMA", "TITAN", "ULTRACEMCO", "WIPRO", "HCLTECH", "NTPC", "POWERGRID",
    "M&M", "TATAMOTORS", "TATASTEEL", "JSWSTEEL", "ASIANPAINT", "ADANIENT",
    "ADANIPORTS", "BAJAJFINSV", "BAJAJ-AUTO", "NESTLEIND", "ONGC", "COALINDIA",
    "GRASIM", "INDUSINDBK", "TECHM", "DIVISLAB", "DRREDDY", "CIPLA",
    "APOLLOHOSP", "EICHERMOT", "HEROMOTOCO", "BRITANNIA", "HINDALCO",
    "SBILIFE", "TATACONSUM", "UPL", "DABUR", "HDFCLIFE", "ICICIPRULI",
    "DLF", "PIDILITIND", "BERGEPAINT", "HAVELLS", "BOSCHLTD", "SIEMENS",
    "ABB", "TATAELXSI", "PERSISTENT", "COFORGE", "LTIM", "POLYCAB",
    "TVSMOTOR", "MOTHERSON", "ASHOKLEY", "AMBUJACEM", "ACC", "SHREECEM",
    "BANKBARODA", "PNB", "CANBK", "UNIONBANK", "INDUSIND", "FEDERALBNK",
    "IDFCFIRSTB", "AUBANK", "YESBANK", "ZOMATO", "PAYTM", "DMART", "TRENT",
    "MCDOWELL-N", "VEDL", "HINDZINC", "SAIL", "JINDALSTEL", "COLPAL",
    "MARICO", "GODREJCP", "PAGEIND", "JUBLFOOD", "DEVYANI", "INDUSTOWER",
    "BHARATFORG", "BHEL", "HAL", "BEL", "DIXON", "VOLTAS", "SUZLON",
    "TATAPOWER", "ADANIGREEN", "ADANITRANS", "GAIL", "IOC", "BPCL", "HPCL",
    "PETRONET", "IGL", "GMRINFRA", "IRCTC", "RVNL", "IREDA", "NHPC", "SJVN",
    "PFC", "RECLTD", "LICHSGFIN", "HDFCAMC", "UTIAMC", "CDSL", "BSE", "MCX",
    "ICICIGI", "SHRIRAMFIN", "CHOLAFIN", "MUTHOOTFIN", "MANAPPURAM",
    "BAJAJHLDNG", "SPR_AUTO", "SPORTKING", "MAHINDCIE", "MINDACORP",
    "CRAFTSMAN", "ACI", "ASTRAL", "ATUL", "BALRAMCHIN", "BANKINDIA",
    "BATAINDIA", "BIOCON", "BLUESTARCO", "CASTROLIND", "CESC", "CGPOWER",
    "CHENNPETRO", "CUB", "CUMMINSIND", "CYIENT", "DALBHARAT", "DEEPAKNTR",
    "DELHIVERY", "ESCORTS", "EXIDEIND", "FINEORG", "GILLETTE", "GLAND",
    "GLAXO", "GODREJIND", "GODREJPROP", "GUJGASLTD", "HAPPSTMNDS",
    "HINDCOPPER", "IDBI", "IDEA", "IIFL", "INDHOTEL", "INDIGO", "JKCEMENT",
    "JSL", "JSWENERGY", "JYOTHYLAB", "KANSAINER", "KAYNES", "KPITTECH",
    "LALPATHLAB", "LAURUSLABS", "LICI", "LUPIN", "MAHABANK", "MAXHEALTH",
    "MFSL", "MGL", "MPHASIS", "MRF", "NATCOPHARM", "NAUKRI", "NAVINFLUOR",
    "NCC", "NIACL", "OBEROIRLTY", "OFSS", "OIL", "PHOENIXLTD", "PIIND",
    "PNBHOUSING", "POLYMED", "PRESTIGE", "RAMCOCEM", "RBLBANK", "REDINGTON",
    "SBICARD", "SHRIRAMPPS", "SIGNATURE", "SONACOMS", "STARHEALTH", "SUNTV",
    "SUPREMEIND", "SWIGGY", "SYNGENE", "TATACOMM", "THERMAX", "TIINDIA",
    "TORNTPHARM", "TORNTPOWER", "UCOBANK", "UNOMINDA", "VBL", "VIPIND",
    "WELCORP", "WESTLIFE", "WHIRLPOOL", "ZYDUSLIFE",
]


# ---------------------------------------------------------------------------
# data: Yahoo Finance (the only source)
# ---------------------------------------------------------------------------
# NSE symbol -> Yahoo symbol (renamed listings)
_YF_ALIASES = {"SPR_AUTO": "SHRIPISTON"}


def _fetch_yfinance(sym: str, start: dt.date, end: dt.date) -> dict | None:
    """Daily bars for ONE symbol from Yahoo Finance; None when no data.

    yfinance history(end=...) is EXCLUSIVE, so +1 day is added (the same
    fix the live scanner's _yf_daily applies) - otherwise the newest bar
    is silently dropped.
    """
    import yfinance as yf
    yf_sym = _YF_ALIASES.get(sym, sym)
    df = yf.Ticker(f"{yf_sym}.NS").history(
        start=start, end=end + dt.timedelta(days=1), auto_adjust=True)
    if df is None or df.empty:
        return None
    df = df.dropna(subset=["Open", "High", "Low", "Close"])
    if df.empty:
        return None
    return {
        "open": df["Open"].to_numpy(float),
        "high": df["High"].to_numpy(float),
        "low": df["Low"].to_numpy(float),
        "close": df["Close"].to_numpy(float),
        "volume": df["Volume"].to_numpy(float),
        "dates": [d.date().isoformat() for d in df.index],
    }


class _Pacer:
    """Serial pacer: enforce a minimum gap between Yahoo calls.

    An un-paced burst of history() calls gets the runner IP 429-rate-
    limited by Yahoo almost immediately (the 2026-08-24 silent-outage
    lesson that produced the live scanner's _YfGate). The backtest loop
    is serial, so a simple monotonic-clock pacer is enough.
    """

    def __init__(self, min_interval: float) -> None:
        self.min_interval = min_interval
        self._last = 0.0

    def wait(self) -> None:
        gap = self.min_interval - (time.monotonic() - self._last)
        if gap > 0:
            time.sleep(gap)
        self._last = time.monotonic()


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------

def rolling_prev_max(a: np.ndarray, w: int) -> np.ndarray:
    """out[i] = max(a[i-w : i]) (max of the w bars BEFORE i); NaN at start."""
    out = np.full(len(a), np.nan)
    if len(a) < 2:
        return out
    from collections import deque
    dq: deque = deque()
    for i in range(len(a)):
        while dq and dq[0] <= i - w - 1:
            dq.popleft()
        if i >= 1:
            while dq and a[dq[-1]] <= a[i - 1]:
                dq.pop()
            dq.append(i - 1)
        if dq:
            out[i] = a[dq[0]]
    return out


def backtest_symbol(sym: str, cfg: ScanConfig, bars: dict,
                    out_rows: list, debug_symbol: str | None = None,
                    since: str | None = None,
                    stats: dict | None = None) -> int:
    """Run the full pattern on every historical day; append signals.
    If debug_symbol == sym, prints WHY candidate days are rejected.

    `since` (ISO date, optional): only REPORT signals dated on/after it.
    The bars handed in may reach further back on purpose - the pattern
    needs min_bars (~26 weeks) of lookback BEFORE the first reportable
    day - but days older than `since` stay unreported (they belong to the
    lookback buffer, not the requested analysis window).

    `stats` (optional dict): per-stage rejection counts so the summary can
    explain WHY a run produced few/zero signals instead of staying silent
    (the 2026-08 silent "0 signals" runs had no way to tell a tight market
    apart from a broken data path)."""
    o = np.asarray(bars["open"], float)
    h = np.asarray(bars["high"], float)
    l = np.asarray(bars["low"], float)
    c = np.asarray(bars["close"], float)
    v = np.asarray(bars.get("volume", np.zeros(len(c))), float)
    dates = bars["dates"]
    n = len(c)
    if n < cfg.min_bars:
        if stats is not None:
            stats["too_few_bars"] = stats.get("too_few_bars", 0) + 1
        return 0

    prev26 = rolling_prev_max(h, cfg.bos_lookback_26w)
    prev45 = rolling_prev_max(h, cfg.bos_lookback_swing)

    def rej(stage: str, sub: str = "") -> None:
        if stats is None:
            return
        stats[stage] = stats.get(stage, 0) + 1
        if sub:
            key = f"{stage}:{sub}"
            stats[key] = stats.get(key, 0) + 1

    found = 0
    dbg = debug_symbol == sym
    # scan EVERY day INCLUDING the last bar: recent signals (e.g. the user's
    # Jul-2026 setups happening days before today) must be included. The very
    # last bar is reported with recent=1 and NaN forward returns (no entry
    # price yet - it is today's still-forming signal).
    for t in range(cfg.min_bars, n):
        # outside the requested analysis window: this day is only lookback
        # (cheap string compare - ISO dates sort lexicographically)
        if since is not None and dates[t] < since:
            continue
        if stats is not None:
            stats["candidates"] = stats.get("candidates", 0) + 1
        # ---------------- BOS (mirrors pattern._find_bos) ----------------
        bos_day, bos_style, brk = None, None, None
        best = None
        for style, prev in (("26w", prev26), ("swing", prev45)):
            for d in range(t - cfg.bos_newest, t - cfg.bos_oldest - 1, -1):
                if d < 0 or np.isnan(prev[d]):
                    continue
                if h[d] > prev[d] * (1 + cfg.bos_break_eps):
                    if best is None or d > best[0]:
                        best = (d, style, float(prev[d]))
                    break
        if best is None:
            if dbg and dates[t] >= "2026-07-20":
                print(f"    [{dates[t]}] no BOS in window")
            rej("bos", "no_bos")
            continue
        bos_day, bos_style, brk = best

        # ------- 26W proximity guard (rejects weak swing breaks) --------
        if bos_style == "swing":
            h26 = prev26[t]
            if np.isnan(h26) or h26 <= 0:
                rej("bos", "swing_no_h26")
                continue
            if np.max(h[bos_day:t + 1]) < h26 * cfg.swing_26w_proximity:
                if dbg and dates[t] >= "2026-07-20":
                    prox = np.max(h[bos_day:t + 1]) / h26 * 100
                    print(f"    [{dates[t]}] swing BOS but peak {prox:.1f}% "
                          f"of 26W high (need {cfg.swing_26w_proximity*100:.0f}%)")
                rej("bos", "swing_proximity")
                continue

        # ------------------------- flush --------------------------------
        pk = bos_day + int(np.argmax(h[bos_day:t + 1]))
        fl = bos_day + int(np.argmin(l[bos_day:t + 1]))
        if fl <= pk:
            rej("flush", "order")
            continue
        peak = float(h[pk])
        flush_low = float(l[fl])
        drop = (peak - flush_low) / peak
        if drop < cfg.flush_min_drop:
            if dbg and dates[t] >= "2026-07-20":
                print(f"    [{dates[t]}] flush drop {drop*100:.1f}% < "
                      f"{cfg.flush_min_drop*100:.0f}%")
            rej("flush", "drop")
            continue
        reds = [(c[i - 1] - c[i]) / c[i - 1] for i in range(pk + 1, t + 1)
                if c[i] < c[i - 1]]
        if not reds or max(reds) < cfg.flush_red_day_min:
            rej("flush", "no_red")
            continue
        if t - fl > cfg.flush_max_age:
            rej("flush", "age")
            continue

        # -------------------------- SSL --------------------------------
        lo = max(0, bos_day - cfg.ssl_pre_lookback)
        ssl = float(np.min(l[lo:bos_day]))
        if ssl <= 0:
            rej("ssl", "none")
            continue
        ratio = flush_low / ssl
        if ratio > 1 + cfg.ssl_tol_up or ratio < 1 - cfg.ssl_tol_dn:
            if dbg and dates[t] >= "2026-07-20":
                print(f"    [{dates[t]}] flush low {flush_low:.1f} vs SSL "
                      f"{ssl:.1f} ({ratio*100:.1f}%) outside tolerance")
            rej("ssl", "tolerance")
            continue

        # ------------------------- hold --------------------------------
        if float(np.min(c[fl:t + 1])) <= ssl:
            rej("hold", "below_ssl")
            continue

        # ---------------------- reversal day ---------------------------
        rng = h[t] - l[t]
        if rng <= 0 or c[t] <= o[t]:
            if dbg and dates[t] >= "2026-07-20":
                print(f"    [{dates[t]}] last candle not green "
                      f"(O {o[t]:.1f} C {c[t]:.1f})")
            rej("candle", "not_green")
            continue
        if (c[t] - o[t]) / rng < cfg.body_ratio_min:
            if dbg and dates[t] >= "2026-07-20":
                print(f"    [{dates[t]}] body ratio {(c[t]-o[t])/rng:.2f} < "
                      f"{cfg.body_ratio_min}")
            rej("candle", "body_ratio")
            continue
        if c[t] < c[t - 1] * (1 + cfg.bounce_min):
            if dbg and dates[t] >= "2026-07-20":
                print(f"    [{dates[t]}] bounce {(c[t]/c[t-1]-1)*100:.1f}% < "
                      f"{cfg.bounce_min*100:.0f}%")
            rej("candle", "bounce")
            continue
        if c[t] < ssl * cfg.near_ssl_close_min:
            rej("candle", "near_ssl")
            continue
        if c[t] >= peak:
            if dbg and dates[t] >= "2026-07-20":
                print(f"    [{dates[t]}] close {c[t]:.1f} >= peak {peak:.1f} "
                      f"(too late)")
            rej("candle", "too_late")
            continue

        # ------------------- prefilter (same as live) ------------------
        # CRITICAL: slice bars to day t ONLY - using the full history would
        # look ahead into the future (close[-1] is the last day of the
        # dataset, not day t) and corrupt every historical signal.
        ok, why = passes_prefilter({k: v[:t + 1] for k, v in bars.items()}, cfg)
        if not ok:
            if debug_symbol and sym == debug_symbol:
                print(f"    [{dates[t]}] prefilter REJECT: {why}")
            if stats is not None:
                for prefix, name in (
                        ("close", "100.html_close"),
                        ("daily", "red_day"),
                        ("RSI", "rsi"),
                        ("MACD", "macd"),
                        ("insufficient", "history"),
                        ("", "other")):
                    if why.startswith(prefix):
                        stats[f"prefilter:{name}"] = \
                            stats.get(f"prefilter:{name}", 0) + 1
                        break
                stats["prefilter"] = stats.get("prefilter", 0) + 1
            continue

        # ------------------- filters (same as live) -------------------
        # min_price: reject penny stocks
        if c[t] < cfg.min_price:
            if dbg and dates[t] >= "2026-07-20":
                print(f"    [{dates[t]}] price {c[t]:.1f} < {cfg.min_price}")
            rej("min_price")
            continue
        # avg_volume: reject illiquid stocks
        # CRITICAL: slice to day t (v[:t+1]) like the prefilter above -
        # the live scanner's volume array ENDS at the signal day, so the
        # 20-day average must too. Using the full series `v` is look-ahead
        # bias: future volume leaks into historical signals (an illiquid
        # stock passes because it became liquid later, and a valid signal
        # is rejected when the stock's volume dries up AFTER day t).
        avg_vol = avg_volume(v[:t + 1], cfg.volume_lookback)
        if avg_vol < cfg.min_avg_volume:
            if dbg and dates[t] >= "2026-07-20":
                print(f"    [{dates[t]}] avg_volume {avg_vol:,.0f} < "
                      f"{cfg.min_avg_volume:,.0f}")
            rej("avg_volume")
            continue

        # --------------------- score (same formula) --------------------
        sig_tmp = {
            "days_since_bos": t - bos_day,
            "flush_drop_pct": drop * 100.0,
            "flush_low": flush_low,
            "ssl": ssl,
            "bounce_pct": (c[t] / c[t - 1] - 1) * 100.0,
            "body_ratio": (c[t] - o[t]) / rng,
        }
        score, _parts = _score(sig_tmp, {"close": c[:t + 1]}, cfg)
        if score < cfg.score_threshold:
            rej("score")
            continue
        # flatten the score components (needed for the last-bar branch too)
        comp = {name: round(v, 1) for name, (v, _m) in (_parts or {}).items()}
        # ------------------- forward returns (next open) ---------------
        if t + 1 >= n:
            # last bar: signal is forming today, no forward data yet
            out_rows.append({
                "symbol": sym, "date": dates[t],
                "close": round(float(c[t]), 2), "score": score,
                "bos": dates[bos_day], "style": bos_style,
                "peak": round(peak, 2), "flush_low": round(flush_low, 2),
                "ssl": round(ssl, 2),
                "score_freshness": comp.get("BOS freshness", ""),
                "score_flush": comp.get("Flush depth", ""),
                "score_ssl": comp.get("SSL precision", ""),
                "score_bounce": comp.get("Reversal bounce", ""),
                "score_body": comp.get("Candle body", ""),
                "score_trend": comp.get("Trend (EMA20/50)", ""),
                "r3": float("nan"), "r5": float("nan"), "r7": float("nan"),
                "r10": float("nan"), "r15": float("nan"),
                "max15": float("nan"), "min15": float("nan"),
                "big_move": 0, "recent": 1,
            })
            found += 1
            continue
        entry = o[t + 1]
        if entry <= 0:
            continue
        fwd = {}
        for k in (3, 5, 7, 10, 15):
            if t + k < n:
                fwd[f"r{k}"] = (c[t + k] / entry - 1) * 100
        # partial windows are fine: recent signals have incomplete forward
        # data (NaN where not enough bars yet) - still reported.
        fwd_win = c[t + 1:min(t + 16, n)]
        max15 = float(np.max(fwd_win) / entry - 1) * 100 if len(fwd_win) else float("nan")
        min15 = float(np.min(fwd_win) / entry - 1) * 100 if len(fwd_win) else float("nan")
        # "big move" = the Sportking-style pop: best close within 15d
        # gains >= 8% from entry (the alert's "big move not fired yet")
        big_move = 1 if (not np.isnan(max15) and max15 >= cfg.big_move_pct) else 0
        # recent = still forming (fewer than 15 forward sessions available)
        recent = 1 if t + 15 >= n else 0

        out_rows.append({
            "symbol": sym, "date": dates[t], "close": round(float(c[t]), 2),
            "score": score, "bos": dates[bos_day], "style": bos_style,
            "peak": round(peak, 2), "flush_low": round(flush_low, 2),
            "ssl": round(ssl, 2),
            "score_freshness": comp.get("BOS freshness", ""),
            "score_flush": comp.get("Flush depth", ""),
            "score_ssl": comp.get("SSL precision", ""),
            "score_bounce": comp.get("Reversal bounce", ""),
            "score_body": comp.get("Candle body", ""),
            "score_trend": comp.get("Trend (EMA20/50)", ""),
            "r3": round(fwd.get("r3", float("nan")), 2),
            "r5": round(fwd.get("r5", float("nan")), 2),
            "r7": round(fwd.get("r7", float("nan")), 2),
            "r10": round(fwd.get("r10", float("nan")), 2),
            "r15": round(fwd.get("r15", float("nan")), 2),
            "max15": round(max15, 2), "min15": round(min15, 2),
            "big_move": big_move, "recent": recent,
        })
        found += 1
    return found


# ---------------------------------------------------------------------------
# stats
# ---------------------------------------------------------------------------

def _stats(rows, key):
    vals = [r[key] for r in rows
            if r.get(key) is not None
            and not (isinstance(r.get(key), float) and np.isnan(r[key]))]
    if not vals:
        return None
    wins = sum(1 for x in vals if x > 0)
    return {
        "n": len(vals), "win": wins / len(vals) * 100,
        "avg": statistics.mean(vals), "med": statistics.median(vals),
    }


def print_summary(rows: list[dict], errors: int, elapsed: float,
                  cfg: ScanConfig) -> None:
    print("\n" + "=" * 74)
    print(f"BACKTEST RESULT — {len(rows)} signals, {errors} fetch errors, "
          f"{elapsed:.0f}s")
    print("=" * 74)
    if not rows:
        print("No signals in the window.")
        return
    print("entry = NEXT day open (alert at close, enter next open)\n")
    labels = {"r3": "3 days", "r5": "5 days", "r7": "7 days",
              "r10": "10 days", "r15": "15 days",
              "max15": "best within 15d", "min15": "worst within 15d"}
    for k in ("r3", "r5", "r7", "r10", "r15", "max15", "min15"):
        s = _stats(rows, k)
        if s:
            print(f"  {labels[k]:>18}: n={s['n']:3d}  win={s['win']:5.1f}%  "
                  f"avg={s['avg']:+6.2f}%  med={s['med']:+6.2f}%")

    # ---- big-move stats (the Sportking-style pop) ----
    bm = sum(1 for r in rows if r.get("big_move"))
    print(f"\n  💥 BIG MOVE (≥ +{cfg.big_move_pct:.0f}% within 15d): "
          f"{bm}/{len(rows)} signals = {bm / len(rows) * 100:.1f}%")

    # ---- score-bucket table: does raising the threshold help? ----
    print("\n  score bucket        n   5d-win   7d-win   5d-avg   big-move%")
    for thr in (50, 55, 60, 65, 70, 75, 80):
        sub = [r for r in rows if r["score"] >= thr]
        if len(sub) < 3:
            continue
        s5 = _stats(sub, "r5")
        s7 = _stats(sub, "r7")
        bm_s = sum(1 for r in sub if r.get("big_move"))
        if s5:
            print(f"  score >= {thr:<3d}       {len(sub):3d}   "
                  f"{s5['win']:5.1f}%   "
                  f"{(s7['win'] if s7 else 0):5.1f}%   "
                  f"{s5['avg']:+6.2f}%   {bm_s / len(sub) * 100:5.1f}%")
    print()


# ---------------------------------------------------------------------------
# Excel export (all scoring details)
# ---------------------------------------------------------------------------

def write_excel(rows: list[dict], path: str) -> None:
    """Write signals to an .xlsx workbook: sheet 'Signals' (all columns)
    + sheet 'Summary' (win-rate stats + score buckets)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        print("openpyxl not installed - skipping Excel export "
              "(pip install openpyxl)", file=sys.stderr)
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Signals"
    fields = list(rows[0].keys())
    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    hdr_font = Font(color="FFFFFF", bold=True)
    ws.append(fields)
    for c in ws[1]:
        c.fill = hdr_fill
        c.font = hdr_font
    for r in rows:
        ws.append([r.get(f, "") for f in fields])
    # freeze header + autofilter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    # column widths
    for i, f in enumerate(fields, 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else "A"].width = \
            max(10, min(20, len(f) + 2))

    # ---- Summary sheet ----
    ws2 = wb.create_sheet("Summary")
    ws2.append(["Backtest summary"])
    ws2["A1"].font = Font(bold=True, size=13)
    ws2.append(["signals", len(rows)])
    labels = {"r3": "3 days", "r5": "5 days", "r7": "7 days",
              "r10": "10 days", "r15": "15 days",
              "max15": "best within 15d", "min15": "worst within 15d"}
    ws2.append([])
    ws2.append(["horizon", "n", "win%", "avg%", "med%"])
    for k in ("r3", "r5", "r7", "r10", "r15", "max15", "min15"):
        s = _stats(rows, k)
        if s:
            ws2.append([labels[k], s["n"], round(s["win"], 1),
                        round(s["avg"], 2), round(s["med"], 2)])
    bm = sum(1 for r in rows if r.get("big_move"))
    ws2.append(["BIG MOVE >=+8%", bm, round(bm / len(rows) * 100, 1) if rows else 0])
    ws2.append([])
    ws2.append(["score bucket", "n", "5d-win%", "7d-win%", "5d-avg%", "big-move%"])
    for thr in (50, 55, 60, 65, 70, 75, 80):
        sub = [r for r in rows if r["score"] >= thr]
        if len(sub) < 3:
            continue
        s5 = _stats(sub, "r5")
        s7 = _stats(sub, "r7")
        bm_s = sum(1 for r in sub if r.get("big_move"))
        if s5:
            ws2.append([f">= {thr}", len(sub), round(s5["win"], 1),
                        round(s7["win"], 1) if s7 else "",
                        round(s5["avg"], 2),
                        round(bm_s / len(sub) * 100, 1)])
    for i, f in enumerate("ABCDEF", 1):
        ws2.column_dimensions[f].width = 12

    wb.save(path)
    print(f"wrote Excel -> {path}")


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------

def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    # Keep the old workflow flags as hidden compatibility options. The
    # backtest was refactored to Yahoo-only, but an already-published GitHub
    # Actions workflow may still pass these flags while it is being updated.
    # Accepting and ignoring them prevents an argparse failure from stopping
    # the run; no Dhan or daily-bar cache is used by this implementation.
    ap.add_argument("--source", choices=["dhan", "yfinance"],
                    default="yfinance", help=argparse.SUPPRESS)
    ap.add_argument("--no-cache", action="store_true", help=argparse.SUPPRESS)
    ap.add_argument("--resume", action="store_true", help=argparse.SUPPRESS)
    ap.add_argument("--symbols-file", default=None)
    ap.add_argument("--years", type=int, default=None,
                    help="years of history to REPORT signals from "
                         "(lookback is fetched on top; overridden by --period)")
    ap.add_argument("--period", choices=["1m", "6m", "1y", "2y", "5y"],
                    default=None,
                    help="preset time period: 1m (1 month), 6m (6 months), "
                         "1y (1 year), 2y (2 years), 5y (5 years). "
                         "Overrides --years and --days.")
    ap.add_argument("--min-score", type=float, default=None)
    ap.add_argument("--out", default="signals_backtest.csv")
    ap.add_argument("--limit", type=int, default=0,
                    help="scan only the first N symbols (0 = all)")
    ap.add_argument("--min-mcap", type=float, default=1000,
                    help="only backtest symbols with market cap >= this many "
                         "crores (uses data/market_cap.csv; 1000 = liquid "
                         "stocks only - matches what the live scanner trades)")
    ap.add_argument("--debug-symbol", default=None,
                    help="print WHY this symbol's candidate days are "
                         "rejected (e.g. --debug-symbol SPORTKING)")
    ap.add_argument("--days", type=int, default=None,
                    help="analyze the last N calendar days (the pattern's "
                         "lookback history is fetched automatically on top; "
                         "overridden by --period)")
    args = ap.parse_args()

    if args.source == "dhan":
        print("NOTICE: --source dhan is deprecated; this backtest uses "
              "Yahoo Finance only.")
    if args.no_cache or args.resume:
        print("NOTICE: cache/resume flags are deprecated; this backtest "
              "does not use a daily-bar cache.")

    cfg = ScanConfig()
    if args.min_score is not None:
        cfg.score_threshold = args.min_score

    # ---- period presets override --years and --days ----
    if args.period:
        period_days = {
            "1m": 30,
            "6m": 180,
            "1y": 365,
            "2y": 730,
            "5y": 1825,
        }
        args.days = period_days[args.period]
        args.years = None  # not used when --days is set
        print(f"period preset: {args.period} = {args.days} days")
    elif args.years is None and args.days is None:
        # default to 5 years if neither --period, --years, nor --days given
        args.years = 5

    if args.symbols_file:
        with open(args.symbols_file) as f:
            symbols = [ln.strip().upper() for ln in f
                       if ln.strip() and not ln.strip().startswith("#")]
    else:
        symbols = list(UNIVERSE)
    if args.min_mcap:
        from prefilter import load_mcap, mcap_filter
        mcap = load_mcap(cfg.mcap_file)
        if mcap:
            before = len(symbols)
            symbols, _dropped = mcap_filter(symbols, mcap, args.min_mcap)
            print(f"mcap filter: {before} -> {len(symbols)} symbols "
                  f"(>= {args.min_mcap:.0f} Cr)")
        else:
            print(f"WARNING: {cfg.mcap_file} not found - min-mcap skipped")
    if args.limit:
        symbols = symbols[:args.limit]
    period_info = f", period={args.period}" if args.period else ""

    # ---- analysis window vs FETCH window --------------------------------
    # --period/--days/--years define the window signals are REPORTED from.
    # The pattern itself needs min_bars (~26 weeks) of history before the
    # first reportable day, so the fetch window extends the analysis window
    # by that lookback. Without the buffer every short preset (1m = ~21
    # bars, 6m = ~124 bars < min_bars=160) fails 100% of symbols as
    # "no data" and can never report a single signal.
    analysis_days = args.days if args.days else 365 * (args.years or 5)
    # min_bars trading days ~= min_bars * 7/5 calendar days (+ margin)
    lookback_days = int(cfg.min_bars * 7 / 5) + 15
    end = dt.date.today()
    since_date = end - dt.timedelta(days=analysis_days)
    since_iso = since_date.isoformat()
    fetch_start = since_date - dt.timedelta(days=lookback_days)

    print(f"source=yfinance universe={len(symbols)} symbols{period_info}, "
          f"{analysis_days / 365:.1f}y window (+{lookback_days}d lookback "
          f"fetched), min_score={cfg.score_threshold}")

    rows: list[dict] = []
    errors = 0
    attempted = 0
    aborted_outage = False
    t0 = time.time()

    stats: dict = {}
    bar_counts: list[int] = []
    fetched_ok = 0

    pacer = _Pacer(cfg.yf_min_interval)
    for i, sym in enumerate(symbols, 1):
        attempted = i
        try:
            pacer.wait()
            bars = _fetch_yfinance(sym, fetch_start, end)
            if (bars is None or len(bars.get("close", [])) < cfg.min_bars) \
                    and cfg.yf_retry_delay > 0:
                # ONE retry after a short pause (transient Yahoo 429s) -
                # same policy as the live scanner's _YfGate
                time.sleep(cfg.yf_retry_delay)
                pacer.wait()
                bars = _fetch_yfinance(sym, fetch_start, end)
            if bars is None or len(bars.get("close", [])) < cfg.min_bars:
                errors += 1
                print(f"  {i}/{len(symbols)} {sym:12s} FAIL "
                      f"(yfinance: no data)", flush=True)
            else:
                fetched_ok += 1
                bar_counts.append(len(bars["close"]))
                got = backtest_symbol(sym, cfg, bars, rows, args.debug_symbol,
                                      since=since_iso, stats=stats)
                print(f"  {i}/{len(symbols)} {sym:12s} "
                      f"bars={len(bars['close']):4d} signals={got}",
                      flush=True)
        except Exception as e:  # noqa: BLE001
            errors += 1
            print(f"  {i}/{len(symbols)} {sym:12s} ERROR {str(e)[:70]}",
                  flush=True)
        # ---- fail fast on a dead data source: if the FIRST
        #      data_outage_min_symbols symbols ALL failed to fetch, Yahoo is
        #      not answering this runner - grinding through the rest (2
        #      paced calls each) only burns 5-10 minutes to reach the same
        #      outage verdict. Abort now and fail the run loudly.
        if attempted >= cfg.data_outage_min_symbols and errors == attempted:
            aborted_outage = True
            print(f"  fetch dead for all {attempted} symbols so far - "
                  f"aborting early (skipping the remaining "
                  f"{len(symbols) - attempted})", flush=True)
            break

    # ---- cooldown: keep only the FIRST signal per symbol within
    #      cooldown_days (backtest-verified: repeats win 30% vs 63%) ----
    if cfg.cooldown_days > 0 and rows:
        rows.sort(key=lambda r: (r["symbol"], r["date"]))
        dedup, last_by_sym = [], {}
        for r in rows:
            sym = r["symbol"]
            if sym in last_by_sym:
                try:
                    gap = (dt.date.fromisoformat(r["date"])
                           - dt.date.fromisoformat(last_by_sym[sym])).days
                    if gap <= cfg.cooldown_days:
                        continue
                except ValueError:
                    pass
            last_by_sym[sym] = r["date"]
            dedup.append(r)
        print(f"cooldown: {len(rows)} -> {len(dedup)} signals "
              f"(dropped {len(rows)-len(dedup)} repeats within "
              f"{cfg.cooldown_days}d)")
        rows = dedup

    rows.sort(key=lambda r: r["date"])
    if rows:
        with open(args.out, "w", newline="") as f:
            w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
            w.writeheader()
            w.writerows(rows)
        print(f"\nwrote {len(rows)} signals -> {args.out}")
        # also write Excel with full scoring details
        xlsx = os.path.splitext(args.out)[0] + ".xlsx"
        write_excel(rows, xlsx)

    print_summary(rows, errors, time.time() - t0, cfg)

    # ---- zero-signal diagnostics: say WHY (the silent-empty fix) ----
    # A run that fetches healthy data but reports 0 signals is a legitimate
    # outcome, but SHOULD say which stage filtered everything out so the
    # next run is not a blind re-run of a broken or mis-tuned pipeline.
    if stats.get("candidates"):
        groups = ("bos", "flush", "ssl", "hold", "candle", "prefilter",
                  "min_price", "avg_volume", "score", "too_few_bars")
        top = sorted(((k, v) for k, v in stats.items()
                      if k in groups or k.startswith("prefilter:")),
                     key=lambda kv: -kv[1])
        detail = " | ".join(f"{k}={v}" for k, v in top)
        if bar_counts:
            bar_counts.sort()
            med = bar_counts[len(bar_counts) // 2]
            fetch = (f"fetched={fetched_ok}/{attempted} "
                     f"bars(med={med},min={bar_counts[0]},max={bar_counts[-1]})")
        else:
            fetch = f"fetched=0/{attempted}"
        print(f"\nREJECTION STATS: candidates={stats.get('candidates', 0)} "
              f"signals={len(rows)} | {detail} | {fetch}")
        if not rows and not outage:
            print(f"::warning title=Backtest 0 signals::0 signals but the data "
                  f"looks healthy ({fetch}). Candidate-day outcomes: {detail}")

    # ---- DATA OUTAGE: most symbols fetched NOTHING (Yahoo unreachable /
    #      rate-limiting this IP). "0 signals" in this state is a lie - the
    #      backtest was blind. Same policy as the live scanner (see
    #      scanner.py / the 2026-08-24 silent-outage postmortem): a loud
    #      ::error:: annotation for the Actions run page plus a non-zero
    #      exit code so the run turns RED instead of a green empty result.
    total = attempted
    outage = (total >= cfg.data_outage_min_symbols
              and errors >= cfg.data_outage_error_frac * total)
    if outage:
        pct = 100.0 * errors / max(total, 1)
        suffix = " (aborted early)" if aborted_outage else ""
        print(f"::error title=Backtest data outage::{errors}/{total} symbols "
              f"failed to fetch ({pct:.0f}%){suffix} - Yahoo Finance "
              f"unreachable or rate-limiting this IP. '0 signals' is "
              f"INVALID for this run: rotate the network/retry later.")
        print(f"ERROR: DATA OUTAGE - {errors}/{total} fetches failed"
              f"{suffix}; results are not a real backtest.", file=sys.stderr)
        return EXIT_DATA_OUTAGE
    return 0


if __name__ == "__main__":
    sys.exit(main())
